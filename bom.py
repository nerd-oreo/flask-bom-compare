from openpyxl import load_workbook

class Bom:
    def __init__(self):
        self.BOM = dict()     # {unique_id:item}
        self.uid_BOM = list() # unique id list for BOM A
        
    def load_xls(self, filename=None, columns=None):
        def load(ws):
            BOM = dict()
            uid_BOM = list()
            parent_stack = list()
            current_parent = None
            last_item = None
            
            min_row = 2
            max_row = ws.max_row
            for row in range(min_row, max_row):
                print("Row: {} - Max Row: {}".format(row, max_row))
                
                level = ws[col_level + str(row)].value
                
                if level != None:
                    # get data from excel sheet
                    level = Helper.convert_level_to_number(str(ws[col_level + str(row)].value))
                    number = str(ws[col_number + str(row)].value)
                    description = str(ws[col_description + str(row)].value)
                    rev = str(ws[col_rev + str(row)].value).split(" ")[0]
                    qty = Helper.convert_qty_to_number(str(ws[col_qty + str(row)].value))
                    
                    item = Item()
                    item.set_item(level, number, description, rev, qty)
                    
                    if ws[col_ref_des + str(row)].value != None:
                        ref_des = str(ws[col_ref_des + str(row)].value)
                        item.set_ref_des(ref_des)
                        
                    if ws[col_mfg_name + str(row)].value != None and ws[col_mfg_number + str(row)].value != None:
                        mfg_name = str(ws[col_mfg_name + str(row)].value)
                        mfg_number = str(ws[col_mfg_number + str(row)].value)
                        item.set_avl(mfg_name, mfg_number)
                    
                    if item.level == 0:                 # top level
                        current_parent = item
                        item.set_parent()
                    elif item.level > last_item.level:  # swapping to a lower level
                        parent_stack.append(current_parent)
                        current_parent = last_item   
                    elif item.level < last_item.level:  # swapping back to upper level
                        for i in range(last_item.level - item.level):
                            current_parent = parent_stack.pop()
                    elif item.level == last_item.level:
                        pass
                    item.set_parent(parent=current_parent)
                    BOM[item.unique_id] = item
                    uid_BOM.append(item.unique_id)
                    last_item = item

                else: # if level is None type, only collect data from column V and X
                    mfg_name = str(ws[col_mfg_name + str(row)].value)
                    mfg_number = str(ws[col_mfg_number + str(row)].value)
                    
                    key = uid_BOM[len(uid_BOM)-1]
                    item = BOM[key]
                    item.set_avl(mfg_name,mfg_number)
                    BOM.update({key:item})

            return BOM, uid_BOM
        
        if len(columns) != 8:
            raise Exception('Number of columns isn\'t matched.')
        else:
            col_level, col_number, col_description, col_qty, col_rev, col_ref_des, col_mfg_name, col_mfg_number = columns

            wb = load_workbook(filename=filename,read_only=False)
            if len(wb.sheetnames) != 2:
                raise Exception('Less or more than two sheets in the template.')
            else:
                self.BOM, self.uid_BOM = load(wb.worksheets[0])
                #self.BOM_B, self.uid_BOM_B = load(wb.worksheets[1])

class Item:
    def __init__(self):
        self.unique_id = ''    # unique id of each item, format: <parent_level>:<parent_number>:<child_level>:<child_number>
        self.level = 0
        self.number = ''
        self.description = ''
        self.rev = ''
        self.quantity = 0
        self.ref_des = list()
        self.avl = list()           # avl list as list of dict
        self.parent = None
        self.change_status = list()
    
    def __repr__(self):
        s = 'Unique Id: {}\n' \
            'Parent: {}\n' \
            'Level: {}\n' \
            'Number: {}\n' \
            'Description: {}\n' \
            'Rev: {}\n' \
            'Qty: {}\n' \
            'Ref_Des: {}\n' \
            'AVL: {}\n'.format(self.unique_id, self.parent.number, self.level, self.number, self.description, self.rev, self.quantity, self.ref_des, self.avl)
        return s
    
    def set_item(self, level, number, description, rev, quantity):
        self.level = level
        self.number = number
        self.description = description.encode('utf-8')
        self.rev = rev
        self.quantity = quantity
    
    def _set_unique_id(self):
        if self.parent == None:
            self.unique_id = '{}:{}'.format(self.level, self.number)
        else:
            self.unique_id = '{}:{}:{}:{}'.format(self.parent.level, self.parent.number, self.level, self.number)
    
    def set_ref_des(self, ref_des):
        if ref_des != None:
            self.ref_des = ref_des.split(',')
        
    def set_avl(self, mfg_name, mfg_number):
        temp = {mfg_name:mfg_number}
        self.avl.append(temp)
        
    def set_parent(self, parent=None):
        self.parent = parent
        self._set_unique_id()

class Helper:
    """ To grouping all utilities functions """
    def convert_qty_to_number(qty):
        if qty == "None":
            return None
        else:
            try:
                return int(qty)
            except ValueError:
                return float(qty)
        
    def convert_level_to_number(level):
        try:
            return int(level)
        except (TypeError, ValueError):
            return None
            
class BomCompare:
    def __init__(self):
        pass