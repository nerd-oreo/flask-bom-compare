from openpyxl import load_workbook
from item import Item
from helper import Helper

filename = "test_template_2.xlsx"

col_level = "A"
col_number = "G"
col_description = "H"
col_qty = "I"
col_rev = "L"
col_ref_des = 'O'
col_mfg_name = "V"
col_mfg_number =  "X"

BOM_A = list()
parent_stack = list()
current_parent = None
last_item = None

wb = load_workbook(filename=filename, read_only=False)
ws = wb['Sheet1']

for row in range(2, ws.max_row+1):
    level = ws[col_level + str(row)].value
    
    if level != None:
        # get data from excel sheet
        level = Helper.convert_level_to_number(str(ws[col_level + str(row)].value))
        number = str(ws[col_number + str(row)].value)
        description = str(ws[col_description + str(row)].value)
        rev = str(ws[col_rev + str(row)].value).split(" ")[0]
        qty = Helper.convert_qty_to_number(str(ws[col_qty + str(row)].value))
        
        item = Item()
        item.set_item(level, number, description, rev, qty)
        
        if ws[col_ref_des + str(row)].value != None:
            ref_des = str(ws[col_ref_des + str(row)].value)
            item.set_ref_des(ref_des)
            
        if ws[col_mfg_name + str(row)].value != None and ws[col_mfg_number + str(row)].value != None:
            mfg_name = str(ws[col_mfg_name + str(row)].value)
            mfg_number = str(ws[col_mfg_number + str(row)].value)
            item.set_avl(mfg_name, mfg_number)
        BOM_A.append(item)


        if item.level == 0:                 # top level
            current_parent = item
            item.set_parent()
        elif item.level > last_item.level:  # swapping to a lower level
            parent_stack.append(current_parent)
            current_parent = last_item   
        elif item.level < last_item.level:  # swapping back to upper level
            current_parent = parent_stack.pop()
        elif item.level == last_item.level:
            pass
        item.set_parent(parent=current_parent)
        last_item = item   

    else: # if level is None type, only collect data from column V and X
        mfg_name = str(ws[col_mfg_name + str(row)].value)
        mfg_number = str(ws[col_mfg_number + str(row)].value)
        
        item = BOM_A[len(BOM_A)-1]
        item.set_avl(mfg_name,mfg_number)
        BOM_A.append(item)


for item in BOM_A:
    print(item)
