from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

ALLOWED_EXTENSIONS = {'xlsx'}


def map_header_to_letter(filename, sheet_name):
    header_list = list()
    header_list.append(('', ''))
    header_list.append(('NO AVL', 'NO AVL'))
    wb = load_workbook(filename=filename, read_only=False)

    ws = wb[sheet_name]
    for i in range(1, ws.max_column+1):
        c = ws.cell(row=1, column=i)
        col_n = get_column_letter(i)
        header_list.append((col_n, c.value))
    return header_list


def get_worksheet_choices(filename):
    wb = load_workbook(filename=filename, read_only=False)
    ws_choices = list()
    ws_choices.append(('', ''))
    for s in wb.sheetnames:
        ws_choices.append((s, s))
    return ws_choices


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def convert_qty_to_number(qty):
    if qty == "None" or qty == '':
        return None
    else:
        try:
            return int(qty)
        except ValueError:
            return float(qty)


def convert_level_to_number(level):
    try:
        return int(level)
    except (TypeError, ValueError):
        return None


def write_to_worksheet(ws, row, column, item, column_change):
    ws[column['level'] + str(row)] = item.level
    ws[column['number'] + str(row)] = item.number
    ws[column['description'] + str(row)] = item.description
    ws[column['rev'] + str(row)] = item.rev
    ws[column['qty'] + str(row)] = item.quantity
    ws[column['ref_des'] + str(row)] = ','.join(item.ref_des)

    x = 'X'
    ws[column_change + str(row)] = x
    return ws


def write_to_worksheet_ref(ws, row, column, item, column_change, column_ref_change):
    ws[column['level'] + str(row)] = item.level
    ws[column['number'] + str(row)] = item.number
    ws[column['description'] + str(row)] = item.description
    ws[column['rev'] + str(row)] = item.rev
    ws[column['qty'] + str(row)] = item.quantity
    ws[column['ref_des'] + str(row)] = ','.join(item.ref_des)

    x = 'X'
    for status in item.change_status:
        if status == 'c_description':
            ws[column_change['description'] + str(row)] = x
        elif status == 'c_rev':
            ws[column_change['rev'] + str(row)] = x
        elif status == 'c_qty':
            ws[column_change['qty'] + str(row)] = x
        elif status == 'c_ref_des':
            ws[column_change['ref_des'] + str(row)] = x
            if column_ref_change is not None:
                if len(item.ref_des_change) == 1:
                    ws[column_ref_change + str(row)] = item.ref_des_change[0]
                elif len(item.ref_des_change):
                    ws[column_ref_change + str(row)] = ','.join(item.ref_des_change)
    return ws


def write_to_worksheet_avl(ws, row, column, item_number, item_mfg_name, item_mfg_number):
    ws[column['number'] + str(row)] = item_number
    ws[column['mfg_name'] + str(row)] = item_mfg_name
    ws[column['mfg_number'] + str(row)] = item_mfg_number
    return ws


def get_value(index, l):
    try:
        return l[index]
    except IndexError:
        None


def get_max(a, b):
    m = len(a)
    if len(b) > m:
        m = len(b)
    return m


def load_clean_name():
    from config import BASE_DIR

    clean_name_path = os.path.join(BASE_DIR, 'CleanName.xlsx')
    clean_name_dict = dict()
    wb = load_workbook(filename=clean_name_path, read_only=False)
    ws = wb['CleanName']
    for i in range(2, ws.max_row+1):
        name = str(ws.cell(row=i, column=1).value).upper()
        clean_name = str(ws.cell(row=i, column=2).value).upper()
        clean_name_dict[name] = clean_name
    return clean_name_dict

