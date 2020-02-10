from openpyxl import load_workbook
from app.utils import convert_level_to_number, convert_qty_to_number
from app.item import Item
from app.header import Header

PARENT = 'parent'
CHILD = 'child'


class Bom:
    def __init__(self):
        self.title = ""
        self.filename = ""
        self.file_path = ""
        self.sheet_name = ""
        self.header = Header()
        self.profile_list = list()
        self.bom = dict()  # {unique_id:item}
        self.uid_bom = list()  # unique id list for bom A

    def set_header_list(self, level, number, description, rev, qty, ref_des, ref_des_delimiter,mfg_name, mfg_number):
        self.header.level = level
        self.header.number = number
        self.header.description = description
        self.header.rev = rev
        self.header.qty = qty
        self.header.ref_des = ref_des
        self.header.ref_des_delimiter = str(ref_des_delimiter)
        self.header.mfg_name = mfg_name
        self.header.mfg_number = mfg_number

    def load_excel(self):
        parent_stack = list()
        current_parent = None
        last_item = None

        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name]

        min_row = 2
        max_row = ws.max_row
        for row in range(min_row, max_row):
            print("Row: {} - Max Row: {}".format(row, max_row))
            level = ws[self.header.level + str(row)].value

            if level is not None:
                # get data from excel sheet
                level = convert_level_to_number(str(ws[self.header.level + str(row)].value))
                number = str(ws[self.header.number + str(row)].value)

                description = str(ws[self.header.description + str(row)].value)
                rev = str(ws[self.header.rev + str(row)].value).split(" ")[0]
                qty = convert_qty_to_number(str(ws[self.header.qty + str(row)].value))

                item = Item()
                item.set_item(level, number, description, rev, qty)

                if ws[self.header.ref_des + str(row)].value is not None:
                    ref_des = str(ws[self.header.ref_des + str(row)].value)
                    delimiter = ''
                    if self.header.ref_des_delimiter == 'COMMA':
                        delimiter = ','
                    elif self.header.ref_des_delimiter == 'SPACE':
                        delimiter = ' '
                    item.set_ref_des(ref_des, delimiter)

                if ws[self.header.mfg_name + str(row)].value is not None and ws[self.header.mfg_number + str(row)].value is not None:
                    mfg_name = str(ws[self.header.mfg_name + str(row)].value)
                    mfg_number = str(ws[self.header.mfg_number + str(row)].value)
                    item.set_avl(mfg_name, mfg_number)

                if item.level == 0:  # top level
                    current_parent = item
                    item.type = PARENT
                    item.set_parent()
                elif item.level > last_item.level:  # swapping to a lower level
                    parent_stack.append(current_parent)
                    last_item.type = PARENT
                    current_parent = last_item
                elif item.level < last_item.level:  # swapping back to upper level
                    item.type = CHILD
                    for i in range(last_item.level - item.level):
                        current_parent = parent_stack.pop()
                elif item.level == last_item.level:
                    item.type = CHILD
                    pass

                item.set_parent(parent=current_parent)
                self.bom[item.unique_id] = item
                self.uid_bom.append(item.unique_id)
                last_item = item

            else:  # if level is None type, only collect data from column V and X
                mfg_name = str(ws[self.header.mfg_name + str(row)].value)
                mfg_number = str(ws[self.header.mfg_number + str(row)].value)

                key = self.uid_bom[len(self.uid_bom) - 1]
                item = self.bom[key]
                item.set_avl(mfg_name, mfg_number)
                self.bom.update({key: item})

    def apply_profile(self):
        for key in self.uid_bom:
            item = self.bom[key]
            for profile in self.profile_list:
                if item.type is profile.type:
                    item.number = profile.apply(item.number)
            self.bom.update({key: item})

    def update(self):
        temp_bom = dict()
        temp_uid_bom = list()
        for key in self.uid_bom:
            item = self.bom[key]
            item.update_unique_id()
            temp_key = item.unique_id
            temp_bom[temp_key] = item
            temp_uid_bom.append(temp_key)
        self.uid_bom = temp_uid_bom
        self.bom = temp_bom
